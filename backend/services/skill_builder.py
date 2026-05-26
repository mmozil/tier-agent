"""Skill builder — converte PDF/Sheet/texto em skill markdown e injeta no container Hermes.

Skill format (agentskills.io padrão usado pelo Hermes):

```markdown
---
name: knowledge-<slug>
description: <1 linha resumo>
version: 1.0
metadata:
  hermes:
    tags: [knowledge, user-uploaded]
---

# <Título>

<conteúdo extraído do arquivo>
```

Skills custom vão pra `~/.hermes/skills/knowledge/` dentro do container Hermes.
"""

import logging
import re
import shlex
from io import BytesIO

from services.container_orchestrator import _ssh_run, container_name

logger = logging.getLogger(__name__)


def _slug(s: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", s.lower().strip())
    return s.strip("-")[:60] or "untitled"


def extract_pdf_text(content: bytes) -> str:
    """Extrai texto de PDF usando pdfplumber."""
    try:
        import pdfplumber
    except ImportError:
        raise RuntimeError("pdfplumber não instalado")
    parts = []
    with pdfplumber.open(BytesIO(content)) as pdf:
        for i, page in enumerate(pdf.pages):
            t = page.extract_text() or ""
            if t.strip():
                parts.append(f"## Página {i+1}\n\n{t.strip()}")
    return "\n\n".join(parts)


def extract_xlsx_text(content: bytes) -> str:
    """Converte Excel em markdown tabelas."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl não instalado")
    wb = load_workbook(BytesIO(content), data_only=True)
    parts = []
    for sheet in wb.worksheets:
        parts.append(f"## {sheet.title}")
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        if header:
            parts.append("| " + " | ".join(str(c or "") for c in header) + " |")
            parts.append("|" + "|".join(["---"] * len(header)) + "|")
        for row in rows[1:]:
            parts.append("| " + " | ".join(str(c if c is not None else "") for c in row) + " |")
    return "\n\n".join(parts)


def build_skill_md(*, title: str, body: str, description: str | None = None) -> str:
    """Gera markdown no formato esperado pelo Hermes."""
    slug = _slug(title)
    desc = description or f"Conhecimento importado: {title}"
    desc = desc.replace("\n", " ")[:200]

    return f"""---
name: knowledge-{slug}
description: {desc}
version: 1.0
metadata:
  hermes:
    tags: [knowledge, user-uploaded]
---

# {title}

{body}
"""


def install_skill_in_container(tenant_id: int, skill_filename: str, skill_md: str) -> str:
    """Escreve skill no ~/.hermes/skills/knowledge/{filename} dentro do container.

    Retorna o path completo dentro do container.
    """
    cname = container_name(tenant_id)
    skills_dir = "/opt/data/.hermes/skills/knowledge"
    file_path = f"{skills_dir}/{skill_filename}"

    # 1) cria dir
    _ssh_run(
        f"docker exec {shlex.quote(cname)} mkdir -p {shlex.quote(skills_dir)} "
        f"&& docker exec {shlex.quote(cname)} chown -R 10000:10000 /opt/data/.hermes/skills"
    )
    # 2) escreve via stdin (base64 pra evitar escape hell)
    import base64

    b64 = base64.b64encode(skill_md.encode()).decode()
    cmd = (
        f"docker exec {shlex.quote(cname)} sh -c "
        f"{shlex.quote(f'echo {b64} | base64 -d > {file_path} && chown 10000:10000 {file_path}')}"
    )
    rc, out, err = _ssh_run(cmd, timeout=30)
    if rc != 0:
        raise RuntimeError(f"Falha escrevendo skill: {err[:200]}")

    logger.info(
        "skill instalada tenant=%s file=%s size=%s bytes",
        tenant_id,
        skill_filename,
        len(skill_md),
    )
    return file_path
